from openpyxl import Workbook
import sys
import os

if len(sys.argv) != 3:
    print("Использование: python script.py <output_name> <input_file>")
    sys.exit(1)

name, res = sys.argv[1], sys.argv[2]

if not os.path.isfile(res):
    print(f"❌ Файл не найден: {os.path.abspath(res)}")
    sys.exit(1)

wb = Workbook()
ws = wb.active

try:
    with open(res, "r", encoding="utf-8") as f:
        lines = f.read().split("\n")  # ← обрабатывает все окончания строк
    print(lines)
    for i, line in enumerate(lines):
        line = line.strip()

        if not line:
            continue
        parts = line.split('\t')
        if i == 0:
            parts = parts[0].split()[::3]
        ws.append(parts)

    output = name + ".xlsx"
    wb.save(output)
    print(f"✅ Успешно! {output} — {ws.max_row} строк, {ws.max_column} колонок")

except Exception as e:
    print(f"💥 Ошибка: {e}")
    raise